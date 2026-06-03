"""Excel .xlsx reader via openpyxl."""

from pathlib import Path

MAX_CHARS = 50_000
MAX_ROWS_PER_SHEET = 200


def read(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if r_idx > MAX_ROWS_PER_SHEET:
                rows.append(f"... (truncated at {MAX_ROWS_PER_SHEET} rows)")
                break
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"--- Sheet: {sheet} ---\n" + "\n".join(rows))
    out = "\n\n".join(parts)
    if len(out) > MAX_CHARS:
        out = out[:MAX_CHARS] + f"\n\n... (truncated at {MAX_CHARS} chars)"
    return out or "(empty workbook)"
